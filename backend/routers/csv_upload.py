import uuid
import io
import csv
import json
from datetime import datetime, timezone
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from models.database import save_csv_rows, get_csv_rows, update_csv_row, delete_csv_upload
from utils.auth import require_admin, require_viewer

router = APIRouter()

KNOWN_HEADERS = {
    "sr. no.": "sr_no", "sr.no.": "sr_no", "sr no": "sr_no", "s.no": "sr_no", "sno": "sr_no",
    "sent by": "sent_by", "sentby": "sent_by",
    "sku name": "sku_name", "sku": "sku_name", "product name": "sku_name",
    "product link": "product_link", "link": "product_link", "url": "product_link",
    "product image": "product_image", "image": "product_image", "image url": "product_image",
    "inquiry sent": "inquiry_sent", "inquiry": "inquiry_sent",
}

def normalize_header(h: str) -> str:
    key = h.strip().lower()
    return KNOWN_HEADERS.get(key, None)


MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB

@router.post("/csv/upload")
async def upload_csv(file: UploadFile = File(...), _admin: str = Depends(require_admin)):
    if not file.filename:
        raise HTTPException(400, "No file provided")

    # Validate extension before reading to give a clear error early
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("csv", "xlsx", "xls"):
        raise HTTPException(400, "Only CSV or Excel files are supported")

    content = await file.read(MAX_UPLOAD_BYTES + 1)
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, "File too large — maximum size is 10 MB")
    upload_id = f"upload_{uuid.uuid4().hex[:10]}"
    upload_name = file.filename
    now = datetime.now(timezone.utc).isoformat()

    rows_to_save = []

    if ext == "csv":
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        raw_rows = list(reader)
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            raw_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                raw_rows.append({headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row)})
        except ImportError:
            raise HTTPException(500, "openpyxl not installed — cannot parse Excel files")

    seq = 1
    for raw in raw_rows:
        mapped: dict = {}
        extra: dict = {}
        for col, val in raw.items():
            norm = normalize_header(col)
            if norm:
                mapped[norm] = str(val).strip() if val else ""
            else:
                extra[col.strip()] = str(val).strip() if val else ""

        # Skip completely empty rows (no SKU name and no product link)
        sku = mapped.get("sku_name", "").strip()
        link = mapped.get("product_link", "").strip()
        if not sku and not link:
            continue

        row_id = f"row_{uuid.uuid4().hex[:10]}"
        rows_to_save.append({
            "id": row_id,
            "upload_id": upload_id,
            "upload_name": upload_name,
            "sr_no": str(seq),          # always sequential, ignores original numbering
            "sent_by": mapped.get("sent_by", ""),
            "sku_name": sku,
            "product_link": link,
            "product_image": mapped.get("product_image", ""),
            "inquiry_sent": mapped.get("inquiry_sent", ""),
            "extra_data": json.dumps(extra),
            "job_id": None,
            "status": "pending",
            "created_at": now,
        })
        seq += 1

    await save_csv_rows(rows_to_save)
    return {"upload_id": upload_id, "upload_name": upload_name, "row_count": len(rows_to_save)}


@router.get("/csv/rows")
async def list_csv_rows(upload_id: str = None, _viewer: str = Depends(require_viewer)):
    rows = await get_csv_rows(upload_id)
    return {"rows": rows}


@router.delete("/csv/upload/{upload_id}")
async def delete_upload(upload_id: str, _admin: str = Depends(require_admin)):
    await delete_csv_upload(upload_id)
    return {"deleted": upload_id}
